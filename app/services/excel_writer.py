from __future__ import annotations
from typing import Dict, Any
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import datetime
import os
import yaml

def _set_cell_value(ws, coord: str, value):
    cell = ws[coord]
    cell.value = value

def _write_by_defined_name(wb, ws_name: str, defined_name: str, value):
    # Resolve defined name -> single cell address
    dn = wb.defined_names.get(defined_name)
    if not dn:
        return False
    # defined name may be a list of destinations
    dest = list(dn.destinations)
    # Use the first destination
    sheet, ext = dest[0]
    ws = wb[sheet]
    # ext can be a range; we just use the top-left cell
    min_col, min_row, max_col, max_row = range_boundaries(ext)
    coord = ws.cell(row=min_row, column=min_col).coordinate
    _set_cell_value(ws, coord, value)
    return True

def fill_excel_template(form_data: Dict[str, Any], mapping_yaml_path: str, template_path: str, out_dir: str = "outputs") -> str:
    with open(mapping_yaml_path, "r", encoding="utf-8") as f:
        mapping = yaml.safe_load(f)

    wb = load_workbook(template_path)
    # default sheet where labels exist (not strictly needed except new names)
    ws_name = wb.sheetnames[0]

    for field, cfg in mapping.get("fields", {}).items():
        dn = cfg.get("named_range")
        fmt = cfg.get("format", "text")
        if dn is None:
            continue
        val = form_data.get(field, None)
        if val is None:
            continue

        # basic formatting
        if fmt == "date":
            # keep as date; if string, leave as-is
            _write_by_defined_name(wb, ws_name, dn, val)
        elif fmt == "currency_jpy":
            try:
                v = int(val)
            except Exception:
                v = val
            _write_by_defined_name(wb, ws_name, dn, v)
        else:
            _write_by_defined_name(wb, ws_name, dn, val)

    # Save
    os.makedirs(out_dir, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(out_dir, f"contract_{ts}.xlsx")
    wb.save(out_path)
    return out_path
